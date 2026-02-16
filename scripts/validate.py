#!/usr/bin/env python3
"""Validate procedures.json against the source Excel file.

Re-reads every data row from each Tab sheet and compares it against the
parsed JSON. Reports mismatches in codes, designations, prices, and
missing/extra rows.

Usage:
    python3 scripts/validate.py [path/to/file.xlsx]
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"

CATEGORY_RE = re.compile(r"RC_\d+ - (.+) - Tab")


def normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h)).strip().upper()


def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 2)
    s = str(val).strip().replace(",", ".")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def find_xlsx_file():
    # Use metadata.json to find the source file that matches data/*.json
    meta_path = DATA_DIR / "metadata.json"
    if meta_path.exists():
        with open(meta_path, encoding="utf-8") as f:
            meta = json.load(f)
        source = REPO_ROOT / meta.get("sourceFile", "")
        if source.exists():
            return source
    # Fallback: first xlsx found
    files = list(REPO_ROOT.glob("*.xlsx"))
    if not files:
        print("ERROR: No .xlsx file found in repo root.", file=sys.stderr)
        sys.exit(1)
    return files[0]


def extract_excel_rows(wb):
    """Read all data rows from the Excel, returning a list of dicts keyed
    by (sheet_name, code, designation) for comparison."""
    excel_rows = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith("- Tab"):
            continue

        match = CATEGORY_RE.match(sheet_name)
        if not match:
            continue

        ws = wb[sheet_name]

        # Get category name from row 2
        category_name = match.group(1).strip()
        for i, row in enumerate(ws.iter_rows(max_row=3, values_only=True), 1):
            if i == 2:
                vals = [v for v in row if v is not None]
                if vals:
                    raw = str(vals[0]).strip()
                    category_name = re.sub(r"^\d+\s*-\s*", "", raw)
                break

        # Find header row and build column map
        col_map = None
        header_row_idx = None

        for row_idx, row in enumerate(ws.iter_rows(), 1):
            cells = list(row)

            if col_map is None:
                for cell in cells:
                    h = normalize_header(cell.value)
                    if h in ("CÓDIGO", "CODIGO", "CÓDIGO "):
                        # Build minimal col_map for validation
                        col_map = {}
                        for idx, c in enumerate(cells):
                            hn = normalize_header(c.value)
                            if hn in ("CÓDIGO", "CODIGO", "CÓDIGO "):
                                col_map["code"] = idx
                            elif hn == "DESIGNAÇÃO":
                                col_map["designation"] = idx
                            elif "ENCARGO" in hn and "ADSE" in hn:
                                col_map["adseCharge"] = idx
                            elif "COPAGAMENTO" in hn:
                                col_map["copayment"] = idx
                        header_row_idx = row_idx
                        break
                continue

            if row_idx <= header_row_idx:
                continue

            code_val = cells[col_map["code"]].value if "code" in col_map else None
            desig_val = cells[col_map["designation"]].value if "designation" in col_map else None

            if code_val is None and desig_val is None:
                continue

            code_num = parse_numeric(code_val)
            if code_num is None:
                continue  # subcategory header or non-data row

            code_str = str(int(code_num))
            designation = str(desig_val).strip() if desig_val else ""
            if not designation:
                continue

            adse_raw = cells[col_map["adseCharge"]].value if "adseCharge" in col_map else None
            copay_raw = cells[col_map["copayment"]].value if "copayment" in col_map else None

            adse_charge = parse_numeric(adse_raw)
            copayment = parse_numeric(copay_raw)

            excel_rows.append({
                "sheet": sheet_name,
                "category": category_name,
                "code": code_str,
                "designation": designation,
                "adseCharge": adse_charge if adse_charge is not None else 0,
                "copayment": copayment,  # None if text like "ver regra 9"
                "copaymentRaw": str(copay_raw).strip() if copay_raw is not None and copayment is None else None,
            })

    return excel_rows


def main():
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1]).resolve()
    else:
        xlsx_path = find_xlsx_file()

    print(f"Validating against: {xlsx_path.name}")
    print(f"JSON source: {DATA_DIR / 'procedures.json'}\n")

    # Load JSON
    with open(DATA_DIR / "procedures.json", encoding="utf-8") as f:
        json_procs = json.load(f)

    # Load Excel
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    excel_rows = extract_excel_rows(wb)
    wb.close()

    # --- Check 1: Row counts per category ---
    print("=" * 60)
    print("CHECK 1: Row counts per category")
    print("=" * 60)

    json_by_cat = {}
    for p in json_procs:
        json_by_cat.setdefault(p["category"], []).append(p)

    excel_by_cat = {}
    for r in excel_rows:
        excel_by_cat.setdefault(r["category"], []).append(r)

    all_cats = sorted(set(list(json_by_cat.keys()) + list(excel_by_cat.keys())))
    count_ok = True
    for cat in all_cats:
        jc = len(json_by_cat.get(cat, []))
        ec = len(excel_by_cat.get(cat, []))
        status = "OK" if jc == ec else "MISMATCH"
        if status == "MISMATCH":
            count_ok = False
        print(f"  {status:10s} {cat}: JSON={jc}, Excel={ec}")

    if count_ok:
        print("\n  All category counts match!\n")
    else:
        print("\n  WARNING: Some category counts differ!\n")

    # --- Check 2: Row-by-row value comparison ---
    print("=" * 60)
    print("CHECK 2: Row-by-row value comparison")
    print("=" * 60)

    # Build lookup: (category, code, designation) -> json proc
    # Use list index for order-based matching since codes can repeat
    json_by_cat_list = {}
    for p in json_procs:
        json_by_cat_list.setdefault(p["category"], []).append(p)

    excel_by_cat_list = {}
    for r in excel_rows:
        excel_by_cat_list.setdefault(r["category"], []).append(r)

    mismatches = []
    missing_in_json = []
    extra_in_json = []

    for cat in all_cats:
        j_list = json_by_cat_list.get(cat, [])
        e_list = excel_by_cat_list.get(cat, [])

        # Build lookup by (code, designation) for matching
        j_lookup = {}
        for p in j_list:
            key = (p["code"], p["designation"])
            j_lookup.setdefault(key, []).append(p)

        e_lookup = {}
        for r in e_list:
            key = (r["code"], r["designation"])
            e_lookup.setdefault(key, []).append(r)

        # Check each Excel row exists in JSON with correct values
        for key, e_entries in e_lookup.items():
            j_entries = j_lookup.get(key, [])
            if not j_entries:
                for e in e_entries:
                    missing_in_json.append(e)
                continue

            # Compare pairwise (by position for duplicates)
            for i, e in enumerate(e_entries):
                if i >= len(j_entries):
                    missing_in_json.append(e)
                    continue

                j = j_entries[i]
                issues = []

                # Compare ADSE charge
                j_adse = round(j["adseCharge"], 2)
                e_adse = round(e["adseCharge"], 2)
                if j_adse != e_adse:
                    issues.append(f"adseCharge: JSON={j_adse}, Excel={e_adse}")

                # Compare copayment
                if e["copayment"] is not None:
                    j_copay = round(j["copayment"], 2)
                    e_copay = round(e["copayment"], 2)
                    if j_copay != e_copay:
                        issues.append(f"copayment: JSON={j_copay}, Excel={e_copay}")
                else:
                    # Non-numeric copayment — check copaymentNote
                    if e["copaymentRaw"] and j.get("copaymentNote") != e["copaymentRaw"]:
                        issues.append(
                            f"copaymentNote: JSON={j.get('copaymentNote')!r}, Excel={e['copaymentRaw']!r}"
                        )

                if issues:
                    mismatches.append({
                        "category": cat,
                        "code": key[0],
                        "designation": key[1][:50],
                        "issues": issues,
                    })

        # Check for extra rows in JSON not in Excel
        for key, j_entries in j_lookup.items():
            e_entries = e_lookup.get(key, [])
            if len(j_entries) > len(e_entries):
                for j in j_entries[len(e_entries):]:
                    extra_in_json.append({
                        "category": cat,
                        "code": j["code"],
                        "designation": j["designation"][:50],
                    })

    if mismatches:
        print(f"\n  MISMATCHES: {len(mismatches)} rows with value differences:\n")
        for m in mismatches[:30]:  # Show first 30
            print(f"    [{m['category']}] Code {m['code']}: {m['designation']}")
            for issue in m["issues"]:
                print(f"      → {issue}")
        if len(mismatches) > 30:
            print(f"    ... and {len(mismatches) - 30} more")
    else:
        print("\n  All values match!")

    if missing_in_json:
        print(f"\n  MISSING from JSON: {len(missing_in_json)} Excel rows not found:\n")
        for m in missing_in_json[:20]:
            print(f"    [{m['category']}] Code {m['code']}: {m['designation'][:50]}")
        if len(missing_in_json) > 20:
            print(f"    ... and {len(missing_in_json) - 20} more")

    if extra_in_json:
        print(f"\n  EXTRA in JSON: {len(extra_in_json)} rows not in Excel:\n")
        for m in extra_in_json[:20]:
            print(f"    [{m['category']}] Code {m['code']}: {m['designation'][:50]}")
        if len(extra_in_json) > 20:
            print(f"    ... and {len(extra_in_json) - 20} more")

    # --- Check 3: Spot checks ---
    print("\n" + "=" * 60)
    print("CHECK 3: Spot checks on known values")
    print("=" * 60)

    spot_checks = [
        ("55010", "TIPAGEM AB0 E RH (D)", 4.17, 1.1),
        ("76032", "CONSUMOS EM SALA CIRÚRGICA (Cirurgia internamento)", 0.8, 0.2),
        ("61851", "CONSULTA ODONTO-ESTOMATOLOGICA", 15.75, 5.25),
        ("10004", "CRÂNIO, UMA INCIDÊNCIA", 8.0, 2.0),
    ]

    json_lookup = {(p["code"], p["designation"]): p for p in json_procs}

    for code, desig, expected_adse, expected_copay in spot_checks:
        key = (code, desig)
        p = json_lookup.get(key)
        if p is None:
            print(f"  FAIL  Code {code} ({desig[:30]}): NOT FOUND in JSON")
            continue

        adse_ok = round(p["adseCharge"], 2) == expected_adse
        copay_ok = round(p["copayment"], 2) == expected_copay

        if adse_ok and copay_ok:
            print(f"  OK    Code {code}: ADSE={p['adseCharge']}, Copag={p['copayment']}")
        else:
            print(f"  FAIL  Code {code}: ADSE={p['adseCharge']} (expected {expected_adse}), "
                  f"Copag={p['copayment']} (expected {expected_copay})")

    # --- Summary ---
    total_issues = len(mismatches) + len(missing_in_json) + len(extra_in_json)
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"  Total Excel rows:   {len(excel_rows)}")
    print(f"  Total JSON rows:    {len(json_procs)}")
    print(f"  Value mismatches:   {len(mismatches)}")
    print(f"  Missing from JSON:  {len(missing_in_json)}")
    print(f"  Extra in JSON:      {len(extra_in_json)}")

    if total_issues == 0:
        print("\n  ✓ VALIDATION PASSED — all data matches the Excel source.")
    else:
        print(f"\n  ✗ VALIDATION FOUND {total_issues} ISSUE(S) — review above.")

    sys.exit(0 if total_issues == 0 else 1)


if __name__ == "__main__":
    main()
