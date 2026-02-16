#!/usr/bin/env python3
"""Parse ADSE Regime Convencionado Excel file into structured JSON.

Usage:
    python3 scripts/parse_excel.py [path/to/file.xlsx]

If no path is given, auto-detects *.xlsx in the repo root.
Outputs: data/procedures.json, data/rules.json, data/metadata.json
"""

import json
import os
import re
import sys
import glob
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"

# Portuguese month mapping for filename date extraction
PT_MONTHS = {
    "janeiro": "01", "fevereiro": "02", "março": "03", "marco": "03",
    "abril": "04", "maio": "05", "junho": "06", "julho": "07",
    "agosto": "08", "setembro": "09", "outubro": "10",
    "novembro": "11", "dezembro": "12",
}

# Canonical category names extracted from sheet names
CATEGORY_RE = re.compile(r"RC_\d+ - (.+) - (Tab|Regras)")


def slugify(text: str) -> str:
    """Convert category name to URL-safe slug."""
    slug = text.lower().strip()
    # Normalize Portuguese characters
    replacements = {"á": "a", "à": "a", "ã": "a", "â": "a",
                    "é": "e", "ê": "e", "í": "i", "ó": "o",
                    "ô": "o", "õ": "o", "ú": "u", "ç": "c"}
    for old, new in replacements.items():
        slug = slug.replace(old, new)
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")


def extract_date_from_filename(filename: str) -> str:
    """Extract date from filename pattern like _01_fevereiro_2026_."""
    pattern = r"_(\d{2})_([a-záàâãéêíóôõúç]+)_(\d{4})_"
    match = re.search(pattern, filename.lower())
    if match:
        day, month_pt, year = match.groups()
        month_num = PT_MONTHS.get(month_pt, "01")
        return f"{year}-{month_num}-{day}"
    return "unknown"


def find_xlsx_file() -> Path:
    """Auto-detect the xlsx file in the repo root."""
    files = list(REPO_ROOT.glob("*.xlsx"))
    if not files:
        print("ERROR: No .xlsx file found in repo root.", file=sys.stderr)
        sys.exit(1)
    if len(files) > 1:
        print(f"WARNING: Multiple .xlsx files found, using: {files[0].name}", file=sys.stderr)
    return files[0]


def get_full_category_name(sheet_title: str) -> str:
    """Map short sheet names to full category names using the sheet's row 2."""
    return sheet_title


def normalize_header(h: str) -> str:
    """Normalize header text for matching."""
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h)).strip().upper()


def build_column_map(header_row):
    """Map column indices to field names based on header text."""
    col_map = {}
    for idx, cell in enumerate(header_row):
        h = normalize_header(cell.value)
        if not h:
            continue
        if h in ("CÓDIGO", "CODIGO", "CÓDIGO "):
            col_map["code"] = idx
        elif h == "DESIGNAÇÃO":
            col_map["designation"] = idx
        elif "ENCARGO" in h and "ADSE" in h:
            col_map["adseCharge"] = idx
        elif "COPAGAMENTO" in h:
            col_map["copayment"] = idx
        elif "QUANT" in h and "MÁX" in h:
            col_map["maxQuantity"] = idx
        elif "QUANTIDADE MÁXIMA" in h:
            col_map["maxQuantity"] = idx
        elif "PRAZO" in h:
            col_map["period"] = idx
        elif "PEQUENA CIRURGIA" in h:
            col_map["smallSurgery"] = idx
        elif "DIAS DE INTERNAMENTO" in h:
            col_map["hospitalizationDays"] = idx
        elif "TIPO DE CÓDIGO" in h or "TIPO DE CODIGO" in h:
            col_map["codeType"] = idx
        elif "OBSERV" in h:
            col_map["observations"] = idx
        elif "NEURONAVEGA" in h:
            col_map["neuronavigation"] = idx
        elif "ROBÓTICA" in h or "ROBOTICA" in h:
            col_map["robotics"] = idx
        elif "LAPAROSCOPIA" in h:
            col_map["laparoscopy"] = idx
        elif "DISPOSITIVOS" in h:
            col_map["medicalDevices"] = idx
        elif "ANESTESIA" in h:
            col_map["anesthesia"] = idx
        elif "COMPONENTES" in h:
            col_map["componentCodes"] = idx
    return col_map


def parse_numeric(val):
    """Safely parse a numeric value, returning None for non-numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 2)
    s = str(val).strip().replace(",", ".")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def parse_tab_sheet(ws, category_name: str, category_slug: str):
    """Parse a Tab sheet into a list of procedure dicts."""
    procedures = []
    col_map = None
    header_row_idx = None
    current_subcategory = None

    for row_idx, row in enumerate(ws.iter_rows(), 1):
        cells = list(row)
        # Find header row (contains CÓDIGO or Código)
        if col_map is None:
            for cell in cells:
                h = normalize_header(cell.value)
                if h in ("CÓDIGO", "CODIGO", "CÓDIGO "):
                    col_map = build_column_map(cells)
                    header_row_idx = row_idx
                    break
            continue

        # Skip rows before data
        if row_idx <= header_row_idx:
            continue

        # Get code value
        code_val = cells[col_map["code"]].value if "code" in col_map else None
        desig_val = cells[col_map["designation"]].value if "designation" in col_map else None

        # Skip empty rows
        if code_val is None and desig_val is None:
            continue

        # Subcategory header: has designation but no numeric code
        if code_val is None and desig_val is not None:
            text = str(desig_val).strip()
            if text and text.upper() != "TABELA":
                current_subcategory = text
            continue

        # Skip non-numeric codes (shouldn't happen but safety check)
        code_num = parse_numeric(code_val)
        if code_num is None:
            # Could be a text-only row
            if desig_val:
                current_subcategory = str(desig_val).strip()
            continue

        code_str = str(int(code_num))
        designation = str(desig_val).strip() if desig_val else ""

        # Skip if no designation
        if not designation:
            continue

        # Parse pricing
        adse_charge = parse_numeric(cells[col_map["adseCharge"]].value) if "adseCharge" in col_map else None
        copayment_raw = cells[col_map["copayment"]].value if "copayment" in col_map else None
        copayment = parse_numeric(copayment_raw)
        copayment_note = None
        if copayment is None and copayment_raw is not None:
            copayment_note = str(copayment_raw).strip()

        proc = {
            "code": code_str,
            "designation": designation,
            "category": category_name,
            "categorySlug": category_slug,
            "adseCharge": adse_charge if adse_charge is not None else 0,
            "copayment": copayment if copayment is not None else 0,
        }

        if current_subcategory:
            proc["subcategory"] = current_subcategory

        if copayment_note:
            proc["copaymentNote"] = copayment_note

        # Optional fields
        if "maxQuantity" in col_map:
            val = parse_numeric(cells[col_map["maxQuantity"]].value)
            if val is not None:
                proc["maxQuantity"] = int(val)

        if "period" in col_map:
            raw = cells[col_map["period"]].value
            if raw is not None:
                num = parse_numeric(raw)
                if num is not None:
                    proc["period"] = f"{int(num)} ano{'s' if num != 1 else ''}"
                else:
                    proc["period"] = str(raw).strip()

        if "hospitalizationDays" in col_map:
            val = parse_numeric(cells[col_map["hospitalizationDays"]].value)
            if val is not None:
                proc["hospitalizationDays"] = int(val)

        if "codeType" in col_map:
            val = cells[col_map["codeType"]].value
            if val is not None:
                proc["codeType"] = str(val).strip()

        if "smallSurgery" in col_map:
            val = cells[col_map["smallSurgery"]].value
            if val is not None:
                proc["smallSurgery"] = str(val).strip().upper() == "SIM"

        if "observations" in col_map:
            val = cells[col_map["observations"]].value
            if val is not None:
                proc["observations"] = str(val).strip()

        procedures.append(proc)

    return procedures


def parse_rules_sheet(ws):
    """Parse a Regras sheet into a list of rule strings."""
    rules = []
    in_rules = False

    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        # Detect the "REGRAS ESPECÍFICAS" header
        first = str(vals[0]).strip().upper()
        if "REGRAS" in first and "ESPECÍFICA" in first:
            in_rules = True
            continue

        if not in_rules:
            continue

        # Rules have a number in col A and text in col B
        if len(vals) >= 2:
            num = vals[0]
            text = str(vals[1]).strip()
            if text and isinstance(num, (int, float)):
                rules.append(text)
            elif text and not isinstance(num, (int, float)):
                # Could be a section header within rules, or a continuation
                # Treat as a rule if it's substantial
                full = str(vals[0]).strip()
                if len(full) > 5:
                    rules.append(full)
        elif len(vals) == 1:
            text = str(vals[0]).strip()
            if text and len(text) > 10:
                rules.append(text)

    return rules


def main():
    # Determine xlsx path
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1]).resolve()
    else:
        xlsx_path = find_xlsx_file()

    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {xlsx_path.name}")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    all_procedures = []
    all_rules = []

    # Process each category (Tab + Regras pair)
    tab_sheets = [n for n in wb.sheetnames if n.endswith("- Tab")]

    for tab_name in tab_sheets:
        match = CATEGORY_RE.match(tab_name)
        if not match:
            print(f"  Skipping unrecognized sheet: {tab_name}")
            continue

        short_name = match.group(1).strip()

        # Get full category name from row 2 of the sheet
        ws_tab = wb[tab_name]
        full_name = short_name
        for i, row in enumerate(ws_tab.iter_rows(max_row=3, values_only=True), 1):
            if i == 2:
                vals = [v for v in row if v is not None]
                if vals:
                    # Strip the leading number like "1 - "
                    raw = str(vals[0]).strip()
                    cleaned = re.sub(r"^\d+\s*-\s*", "", raw)
                    full_name = cleaned
                break

        slug = slugify(full_name)

        # Parse tab
        ws_tab = wb[tab_name]
        procedures = parse_tab_sheet(ws_tab, full_name, slug)
        all_procedures.extend(procedures)
        print(f"  {full_name}: {len(procedures)} procedures")

        # Parse rules
        rules_name = tab_name.replace("- Tab", "- Regras")
        if rules_name in wb.sheetnames:
            ws_rules = wb[rules_name]
            rules = parse_rules_sheet(ws_rules)
            if rules:
                all_rules.append({
                    "category": full_name,
                    "slug": slug,
                    "rules": rules,
                })
                print(f"    → {len(rules)} rules")

    wb.close()

    # Build metadata
    table_date = extract_date_from_filename(xlsx_path.name)
    category_counts = {}
    for p in all_procedures:
        key = (p["category"], p["categorySlug"])
        category_counts[key] = category_counts.get(key, 0) + 1

    metadata = {
        "sourceFile": xlsx_path.name,
        "tableDate": table_date,
        "parsedAt": datetime.now(timezone.utc).isoformat(),
        "totalProcedures": len(all_procedures),
        "categories": [
            {"name": name, "slug": slug, "count": count}
            for (name, slug), count in category_counts.items()
        ],
    }

    # Write output
    DATA_DIR.mkdir(exist_ok=True)

    with open(DATA_DIR / "procedures.json", "w", encoding="utf-8") as f:
        json.dump(all_procedures, f, ensure_ascii=False, indent=2)

    with open(DATA_DIR / "rules.json", "w", encoding="utf-8") as f:
        json.dump(all_rules, f, ensure_ascii=False, indent=2)

    with open(DATA_DIR / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    print(f"\nDone! {len(all_procedures)} procedures across {len(category_counts)} categories.")
    print(f"Table date: {table_date}")
    print(f"Output: {DATA_DIR}/")


if __name__ == "__main__":
    main()
